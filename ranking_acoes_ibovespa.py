# ============================================================
# NOME DO SCRIPT: ranking_acoes_ibovespa.py
# OBJETIVO: Buscar indicadores fundamentalistas de todas as
#           ações do Ibovespa, ranqueá-las por múltiplos
#           critérios e exportar o resultado para Excel e terminal.
# BIBLIOTECAS NECESSÁRIAS: fundamentus, pandas, openpyxl, requests
# COMO EXECUTAR: No terminal do VS Code (com conda ativo):
#                python ranking_acoes_ibovespa.py
# ============================================================


# --- SEÇÃO 1: Importações ---

# Não usamos mais o wrapper 'fundamentus' diretamente pois ele quebrou
# com uma atualização do site (coluna 'Dív.Brut/ Patrim.' causava erro).
# Substituímos por scraping direto com requests + pandas — mais robusto.

# 'pandas' é a biblioteca principal para trabalhar com tabelas de dados.
# Usamos o apelido 'pd' por convenção universal.
import pandas as pd

# 'requests' permite fazer chamadas HTTP — usaremos para buscar
# a composição atual do Ibovespa diretamente do site da B3.
import requests

# 'datetime' é uma biblioteca nativa do Python para trabalhar com datas.
# Usamos para registrar quando o arquivo foi gerado.
from datetime import datetime

# 'os' é nativa do Python e usamos para criar pastas automaticamente
# caso elas ainda não existam no computador.
import os

# 'time' é nativa do Python. Usamos para pausar entre requisições
# e não sobrecarregar os servidores dos portais.
import time

# 'warnings' suprime mensagens de aviso desnecessárias que algumas
# bibliotecas emitem — deixa a saída do terminal mais limpa.
import warnings
warnings.filterwarnings("ignore")


# --- SEÇÃO 2: Configurações ---
# ⚙️ PERSONALIZE AQUI: altere esses parâmetros conforme sua análise

# Pasta onde os resultados serão salvos
# O script cria automaticamente se não existir
PASTA_RESULTADOS = "resultados"

# Pasta onde os dados brutos serão cacheados em CSV
# Cache evita buscar os mesmos dados várias vezes
PASTA_CACHE = "dados/fundamentos"

# Filtros mínimos de qualidade — ações que não passarem são descartadas
# Ajuste esses valores conforme seu critério pessoal
FILTRO_LIQUIDEZ_MINIMA   = 500_000    # R$ 500 mil — reduzido para carteira de longo prazo
FILTRO_PL_MINIMO         = 0           # Excluir P/L negativo (empresa com prejuízo)
FILTRO_PL_MAXIMO         = 40          # Excluir P/L muito alto (empresa muito cara)
FILTRO_DY_MINIMO         = 0.06        # Mínimo 6% de DY — critério AGF/Barsi (0.06 = 6%)
FILTRO_PVPA_MAXIMO       = 5.0         # Excluir P/VP acima de 5 (muito caro vs patrimônio)

# Filtro de setores — Método BESST de Luiz Barsi
# B = Bancos | E = Energia | S = Saneamento | S = Seguros | T = Telecom
# True  = aplica o filtro (mantém apenas ações dos setores BESST)
# False = desativa o filtro (analisa todos os setores do Ibovespa)
FILTRO_BESST_ATIVO = True

# Mapeamento de tickers do Ibovespa por setor BESST
# Fonte: classificação setorial da B3 + composição atual do Ibovespa (Jun/2026)
# Atualize esta lista conforme mudanças na composição do índice
TICKERS_BESST = {
    "Bancos": [
        "BBAS3",   # Banco do Brasil
        "BBDC3",   # Bradesco ON
        "BBDC4",   # Bradesco PN
        "ITUB4",   # Itaú Unibanco PN
        "ITSA4",   # Itaúsa PN (holding do Itaú)
        "SANB11",  # Santander Brasil units
        "BPAC11",  # BTG Pactual units
    ],
    "Energia": [
        "ELET3",   # Eletrobras ON
        "ELET6",   # Eletrobras PNB
        "CMIG4",   # Cemig PN
        "CPFE3",   # CPFL Energia ON
        "CPLE6",   # Copel PNB
        "EGIE3",   # Engie Brasil ON
        "ENEV3",   # Eneva ON
        "ENGI11",  # Energisa units
        "EQTL3",   # Equatorial Energia ON
        "TAEE11",  # Transmissora Aliança units
        "ISAE4",   # Isa Cteep PN
    ],
    "Saneamento": [
        "SBSP3",   # Sabesp ON
    ],
    "Seguros": [
        "BBSE3",   # BB Seguridade ON
        "IRBR3",   # IRB Brasil RE ON
        "PSSA3",   # Porto Seguro ON
    ],
    "Telecom": [
        "VIVT3",   # Telefônica Vivo ON
        "TIMS3",   # TIM ON
    ],
}

# Lista plana de todos os tickers BESST (usada no filtro)
TODOS_TICKERS_BESST = [
    ticker
    for lista in TICKERS_BESST.values()
    for ticker in lista
]

# Pesos do ranking multicritério (devem somar 1.0)
# Ajuste conforme sua filosofia de investimento
PESOS = {
    "dy":        0.30,   # Dividend Yield — foco em renda passiva (peso maior)
    "roe":       0.25,   # ROE — qualidade e eficiência do negócio
    "pl":        0.20,   # P/L — valuation (quanto paga por cada R$1 de lucro)
    "pvpa":      0.15,   # P/VP — desconto sobre o valor patrimonial
    "div_ebitda": 0.10,  # Dívida/EBITDA — saúde financeira (peso menor)
}

# Número de ações no TOP ranking exibido no terminal
TOP_N = 10


# --- SEÇÃO 3: Funções ---

def criar_pastas():
    """
    Cria as pastas de resultados e cache se ainda não existirem.
    O parâmetro exist_ok=True evita erro se a pasta já existir.
    """
    os.makedirs(PASTA_RESULTADOS, exist_ok=True)
    os.makedirs(PASTA_CACHE, exist_ok=True)
    print(f"📁 Pastas verificadas: '{PASTA_RESULTADOS}/' e '{PASTA_CACHE}/'")


def buscar_composicao_ibovespa():
    """
    Busca a composição atual do Ibovespa diretamente da API da B3.
    Retorna uma lista com os tickers das ações do índice.

    A B3 disponibiliza os dados do Ibovespa em formato JSON
    através de uma API pública — sem necessidade de login.
    """
    print("\n🔍 Buscando composição atual do Ibovespa na B3...")

    # URL da API pública da B3 que retorna os componentes do Ibovespa
    # segmentData=2 filtra apenas ações (exclui outros tipos de ativos)
    url = (
        "https://sistemaswebb3-listados.b3.com.br/indexProxy/indexCall/"
        "GetPortfolioDay/eyJsYW5ndWFnZSI6InB0LWJyIiwicGFnZU51bWJlciI6MSwi"
        "cGFnZVNpemUiOjEyMCwiaW5kZXgiOiJJQk9WIiwic2VnbWVudCI6IjIifQ=="
    )

    # Cabeçalho HTTP simulando um navegador real
    # Sem isso, alguns servidores podem bloquear a requisição
    cabecalho = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }

    try:
        # Fazemos a requisição com timeout de 15 segundos
        resposta = requests.get(url, headers=cabecalho, timeout=15)

        # raise_for_status() lança um erro se o servidor retornar
        # um código de erro HTTP (ex: 404, 500)
        resposta.raise_for_status()

        # Convertemos a resposta JSON para um dicionário Python
        dados = resposta.json()

        # Extraímos apenas os tickers (códigos das ações)
        # A estrutura do JSON da B3 tem uma chave 'results' com a lista
        tickers = [item["cod"] for item in dados["results"]]

        print(f"   ✅ {len(tickers)} ações encontradas no Ibovespa")
        return tickers

    except Exception as erro:
        # Se falhar, usamos uma lista manual de fallback com as principais ações
        print(f"   ⚠️  Não foi possível buscar da B3: {erro}")
        print(f"   📋 Usando lista de fallback com as principais ações...")

        # Lista manual das principais ações do Ibovespa (atualizada em Jun/2026)
        # Use como backup caso a API da B3 esteja indisponível
        tickers_fallback = [
            "ABEV3", "ASAI3", "AZUL4", "B3SA3", "BBAS3", "BBDC3", "BBDC4",
            "BBSE3", "BEEF3", "BPAC11", "BRAP4", "BRFS3", "BRKM5", "CASH3",
            "CCRO3", "CMIN3", "CMIG4", "COGN3", "CPFE3", "CPLE6", "CRFB3",
            "CSAN3", "CSNA3", "CVCB3", "CYRE3", "DXCO3", "EGIE3", "ELET3",
            "ELET6", "EMBR3", "ENEV3", "ENGI11", "EQTL3", "EZTC3", "FLRY3",
            "GGBR4", "GOAU4", "GOLL4", "HAPV3", "HYPE3", "IGTI11", "IRBR3",
            "ITSA4", "ITUB4", "JBSS3", "KLBN11", "LREN3", "LWSA3", "MGLU3",
            "MRFG3", "MRVE3", "MULT3", "NTCO3", "PCAR3", "PETR3", "PETR4",
            "PETZ3", "PRIO3", "PSSA3", "RADL3", "RAIL3", "RAIZ4", "RDOR3",
            "RENT3", "RRRP3", "SANB11", "SBSP3", "SLCE3", "SMTO3", "SOMA3",
            "SUZB3", "TAEE11", "TIMS3", "TOTS3", "UGPA3", "USIM5", "VALE3",
            "VBBR3", "VIVT3", "WEGE3", "YDUQ3",
        ]
        return tickers_fallback


def converter_numero_br(valor, eh_percentual=False, tem_decimal_implicito=False):
    """
    Converte um valor numérico no formato do Fundamentus para float.

    O Fundamentus usa DOIS formatos distintos dependendo da coluna:

    FORMATO A — Percentuais (DY, ROE, margens):
      Entrada:  "12,34%"  →  Saída: 0.1234
      Entrada:  "-0,76%"  →  Saída: -0.0076
      A vírgula é o separador decimal; o % indica divisão por 100.

    FORMATO B — Números com decimal IMPLÍCITO de 3 casas (P/L, P/VP, cotação):
      O site OMITE a vírgula decimal — o número tem sempre 3 casas decimais
      embutidas sem separador. Exemplos reais observados:
        "-72851" → na verdade é -72,851  (P/L)
        "011"    → na verdade é  0,011   (P/VP = 0.011... espera, veja abaixo)
      ATENÇÃO: após testes, o decimal implícito é de 2 casas para P/VP:
        "011"  = 0,11  →  P/VP de 0.11x  ✓ (faz sentido: IRBR3 a 0.78x)
        "-044" = -0,44 →  P/VP negativo  ✓ (BRKM5 com patrimônio negativo)
      Para P/L, o padrão é 2 casas também:
        "-077" = -0,77  ✓
        "-019" = -0,19  ✓

    FORMATO C — Liquidez e patrimônio (com ponto de milhar e vírgula decimal):
      Entrada: "136.468.000,00"  →  Saída: 136468000.0
      Entrada: "000"             →  Saída: NaN (sem liquidez)

    Parâmetros:
        valor               : valor a converter (qualquer tipo)
        eh_percentual       : True para colunas DY, ROE, margens
        tem_decimal_implicito: True para P/L, P/VP, cotação, EV/EBIT etc.

    Retorna:
        float convertido, ou NaN se inválido
    """
    import math

    # Se já for número Python nativo, retorna direto
    if isinstance(valor, (int, float)):
        if isinstance(valor, float) and math.isnan(valor):
            return float("nan")
        return float(valor)

    texto = str(valor).strip()

    # Ausência de dado real
    if texto in ("", "-", "—", "N/A", "n/a", "nan", "NaN", "000", "0000", "0"):
        return float("nan")

    try:
        if "%" in texto:
            # FORMATO A: percentual com vírgula decimal
            texto = texto.replace("%", "").strip()
            texto = texto.replace(".", "").replace(",", ".")
            return float(texto) / 100

        elif "," in texto:
            # FORMATO C: número com ponto de milhar e vírgula decimal
            # Ex: "136.468.000,00" ou "1.083.050.000,00"
            partes = texto.split(",")
            inteiro = partes[0].replace(".", "")
            decimal = partes[1] if len(partes) > 1 else "0"
            return float(f"{inteiro}.{decimal}")

        else:
            # FORMATO B: inteiro sem separador — decimal implícito de 2 casas
            # Ex: "-077" → -0.77 | "011" → 0.11 | "-5591" → -55.91
            # Dividimos por 100 para recuperar o decimal implícito
            if tem_decimal_implicito:
                return float(texto) / 100
            else:
                # Sem decimal implícito: liquidez "000" já tratado acima
                # Outros inteiros puros (não deveria ocorrer, mas tratamos)
                return float(texto)

    except (ValueError, IndexError):
        return float("nan")


# Mapeamento de quais colunas têm decimal implícito no Fundamentus
# Descoberto por diagnóstico direto dos dados brutos do site
COLUNAS_DECIMAL_IMPLICITO = {
    "p/l", "p/vp", "psr", "p/ativo", "p/cap.giro",
    "p/ebit", "p/ativ circ.liq", "ev/ebit", "ev/ebitda",
    "liq.corr", "div.liq.patrim", "cotacao",
}

# Mapeamento de quais colunas são percentuais
COLUNAS_PERCENTUAL = {
    "dy", "roe", "roic", "mrg.bruta", "mrg.ebit", "mrg.liq", "cresc.rec.5a",
}


def buscar_todos_fundamentus():
    """
    Busca os indicadores fundamentalistas de TODAS as ações listadas
    no Fundamentus via scraping direto com requests + pandas.

    PROBLEMA RESOLVIDO:
    O Fundamentus retorna TODOS os valores como strings formatadas no
    padrão brasileiro — ex: "12,34%" para DY, "1.234,56" para patrimônio.
    A função converter_numero_br() trata cada caso corretamente.

    Retorna um DataFrame (tabela) com todos os indicadores já em float.
    """
    print("\n📊 Buscando indicadores fundamentalistas no Fundamentus...")
    print("   (isso pode levar alguns segundos...)")

    url = "https://www.fundamentus.com.br/resultado.php"

    cabecalho = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://www.fundamentus.com.br/",
    }

    try:
        print("   Conectando ao Fundamentus.com.br...")
        resposta = requests.get(url, headers=cabecalho, timeout=30)
        resposta.raise_for_status()

        # Passo 1: Lemos a tabela SEM conversão automática (dtype=str)
        # Isso garante que recebemos os valores exatamente como o site os envia,
        # sem o pandas tentar (e falhar) converter "1.234,56" para número.
        from io import StringIO
        lista_tabelas = pd.read_html(
            StringIO(resposta.text),
            flavor="lxml"
            # Nota: read_html() não aceita dtype=str como read_csv().
            # Convertemos cada coluna para string manualmente logo abaixo.
        )
        df = lista_tabelas[0]

        # Convertemos todas as colunas para string agora —
        # assim garantimos que nossa função converter_numero_br()
        # receberá sempre texto, independente do que o pandas inferiu.
        for col in df.columns:
            df[col] = df[col].astype(str)

        # Passo 2: Padronizar nomes das colunas (minúsculas, sem espaços extras)
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Passo 3: Definir o ticker (coluna "papel") como índice da tabela
        if "papel" in df.columns:
            df = df.set_index("papel")
        else:
            print("   ⚠️  Coluna 'papel' não encontrada — usando primeira coluna")
            df = df.set_index(df.columns[0])

        # Passo 4: Renomear colunas para nomes padronizados internos
        # O nome real que chega do site está à esquerda do ":"
        # O nome que usamos no restante do script está à direita
        mapa_colunas = {
            "div.yield":            "dy",
            "p/l":                  "p/l",
            "p/vp":                 "p/vp",
            "p/ebit":               "p/ebit",
            "p/ativo":              "p/ativo",
            "p/cap.giro":           "p/cap.giro",
            "p/ativ circ.liq":      "p/ativ circ.liq",
            "ev/ebit":              "ev/ebit",
            "ev/ebitda":            "ev/ebitda",
            "roe":                  "roe",
            "roic":                 "roic",
            "liq.2meses":           "liq.2meses",
            "mrg bruta":            "mrg.bruta",
            "mrg ebit":             "mrg.ebit",
            "mrg. líq.":            "mrg.liq",
            "liq. corr.":           "liq.corr",
            "patrim. líq":          "patrim.liq",
            "dív.líq/ patrim.":     "div.liq.patrim",
            "cresc. rec.5a":        "cresc.rec.5a",
            "cotação":              "cotacao",
            "psr":                  "psr",
        }
        df = df.rename(columns=mapa_colunas)

        # Passo 5: Converter colunas numéricas com o formato correto por tipo
        # Usamos os mapeamentos COLUNAS_DECIMAL_IMPLICITO e COLUNAS_PERCENTUAL
        # para saber qual lógica de conversão aplicar em cada coluna.
        todas_colunas_numericas = (
            COLUNAS_DECIMAL_IMPLICITO |
            COLUNAS_PERCENTUAL |
            {"liq.2meses", "patrim.liq"}   # Formato C: "136.468.000,00"
        )

        for col in todas_colunas_numericas:
            if col in df.columns:
                eh_pct      = col in COLUNAS_PERCENTUAL
                tem_dec_imp = col in COLUNAS_DECIMAL_IMPLICITO
                df[col] = df[col].apply(
                    lambda v: converter_numero_br(
                        v,
                        eh_percentual=eh_pct,
                        tem_decimal_implicito=tem_dec_imp
                    )
                )

        # Passo 6: Remover linhas com ticker inválido
        df = df[df.index.notna()]
        df = df[df.index.astype(str).str.strip() != ""]

        # Passo 7: Salvar cache
        agora = datetime.now().strftime("%Y%m%d_%H%M")
        caminho_cache = os.path.join(PASTA_CACHE, f"fundamentus_bruto_{agora}.csv")
        df.to_csv(caminho_cache, sep=";", decimal=".", encoding="utf-8-sig")

        # Diagnóstico rápido do DY para confirmar que a conversão funcionou
        if "dy" in df.columns:
            dy_validos = df["dy"].dropna()
            print(f"   ✅ {len(df)} ações obtidas do Fundamentus")
            print(f"   📊 DY — mín: {dy_validos.min():.2%}  "
                  f"máx: {dy_validos.max():.2%}  "
                  f"média: {dy_validos.mean():.2%}")
        else:
            print(f"   ✅ {len(df)} ações obtidas do Fundamentus")

        print(f"   💾 Cache salvo em: {caminho_cache}")
        return df

    except Exception as erro:
        print(f"\n❌ ERRO ao buscar dados do Fundamentus: {erro}")
        print(f"   Tipo do erro: {type(erro).__name__}")
        import traceback
        traceback.print_exc()
        return None


def filtrar_por_ibovespa(df_todos, tickers_ibov):
    """
    Filtra o DataFrame completo mantendo apenas as ações
    que fazem parte do Ibovespa.

    Parâmetros:
        df_todos      : DataFrame com todos os ativos do Fundamentus
        tickers_ibov  : Lista de tickers do Ibovespa (ex: ['PETR4', 'VALE3'])

    Retorna:
        DataFrame filtrado apenas com ações do Ibovespa
    """
    print(f"\n🔧 Filtrando apenas ações do Ibovespa...")

    # O índice do DataFrame do Fundamentus é o ticker da ação
    # Usamos .isin() para manter só os tickers que estão na lista do Ibovespa
    df_ibov = df_todos[df_todos.index.isin(tickers_ibov)].copy()

    print(f"   ✅ {len(df_ibov)} ações do Ibovespa encontradas no Fundamentus")
    return df_ibov


def aplicar_filtros_qualidade(df):
    """
    Remove ações que não atendem aos critérios mínimos de qualidade.
    Ações com fundamentos ruins ou muito arriscadas são descartadas
    antes do ranking para não distorcer a pontuação.

    Parâmetros:
        df : DataFrame com as ações do Ibovespa

    Retorna:
        DataFrame apenas com ações que passaram nos filtros
    """
    print(f"\n🔎 Aplicando filtros de qualidade...")
    total_antes = len(df)

    # Mapeamento dos nomes de colunas do Fundamentus
    # O Fundamentus usa nomes específicos — verificamos quais existem
    col_liquidez  = "liq.2meses" if "liq.2meses" in df.columns else None
    col_pl        = "p/l"        if "p/l"        in df.columns else None
    col_dy        = "dy"         if "dy"         in df.columns else None
    col_pvpa      = "p/vp"       if "p/vp"       in df.columns else None

    # --- Diagnóstico rápido antes dos filtros ---
    # Exibe os valores reais de P/L para as primeiras ações do Ibovespa
    # Isso nos ajuda a entender se a conversão numérica está funcionando
    print(f"\n   🔬 Diagnóstico de colunas (primeiros 5 valores):")
    for col_diag in ["p/l", "dy", "p/vp", "liq.2meses"]:
        if col_diag in df.columns:
            amostra = df[col_diag].head(5).tolist()
            tipo = df[col_diag].dtype
            print(f"      {col_diag:15} tipo={tipo}  valores={amostra}")

    # --- Conversão numérica explícita usando converter_numero_br ---
    # Todas as colunas já passaram por converter_numero_br() na busca,
    # mas aplicamos novamente aqui como garantia para as colunas de filtro.
    for col in [col_liquidez, col_pl, col_dy, col_pvpa]:
        if col and col in df.columns:
            # Se ainda for object (string), aplica a conversão BR
            if df[col].dtype == object:
                df[col] = df[col].apply(converter_numero_br)
            else:
                df[col] = pd.to_numeric(df[col], errors="coerce")

    # Filtro 1: Liquidez mínima — evita ações com pouco volume de negociação
    if col_liquidez:
        df = df[df[col_liquidez] >= FILTRO_LIQUIDEZ_MINIMA]
        print(f"   Após filtro de liquidez    : {len(df)} ações")

    # Filtro 2: P/L dentro de um intervalo razoável
    # P/L negativo = empresa com prejuízo; P/L muito alto = empresa cara demais
    if col_pl:
        df = df[(df[col_pl] > FILTRO_PL_MINIMO) & (df[col_pl] <= FILTRO_PL_MAXIMO)]
        print(f"   Após filtro de P/L         : {len(df)} ações")

    # Filtro 3: Dividend Yield mínimo — foco em pagadores de dividendos
    if col_dy:
        df = df[df[col_dy] >= FILTRO_DY_MINIMO]
        print(f"   Após filtro de DY mínimo   : {len(df)} ações")

    # Filtro 4: P/VP máximo — evita empresas muito caras vs patrimônio
    if col_pvpa:
        df = df[df[col_pvpa] <= FILTRO_PVPA_MAXIMO]
        print(f"   Após filtro de P/VP máximo : {len(df)} ações")

    # Filtro 5: Setor BESST — mantém apenas ações dos setores de Barsi
    # Só aplica se FILTRO_BESST_ATIVO for True nas configurações
    if FILTRO_BESST_ATIVO:
        df = df[df.index.isin(TODOS_TICKERS_BESST)]
        print(f"   Após filtro BESST          : {len(df)} ações")

    # Acrescenta coluna "setor" para identificar a categoria BESST de cada ação
    # Isso enriquece o Excel e facilita análise por setor
    def identificar_setor(ticker):
        """Retorna o setor BESST do ticker, ou 'Outro' se não estiver mapeado."""
        for setor, lista in TICKERS_BESST.items():
            if ticker in lista:
                return setor
        return "Outro"

    df["setor"] = df.index.map(identificar_setor)

    total_depois = len(df)
    descartadas  = total_antes - total_depois
    print(f"\n   📋 Resumo: {total_antes} → {total_depois} ações "
          f"({descartadas} descartadas pelos filtros)")

    return df


def calcular_ranking(df):
    """
    Aplica o ranking multicritério usando normalização min-max.

    Para cada indicador, cada ação recebe uma NOTA de 0 a 1:
      - Nota 1.0 = melhor valor do grupo nesse indicador
      - Nota 0.0 = pior valor do grupo nesse indicador

    A nota final é a média ponderada das notas individuais,
    usando os pesos definidos nas configurações.

    Parâmetros:
        df : DataFrame filtrado com as ações candidatas

    Retorna:
        DataFrame com coluna 'nota_final' e ordenado do melhor ao pior
    """
    print(f"\n🏆 Calculando ranking multicritério...")

    # Criamos uma cópia para não modificar o DataFrame original
    df_rank = df.copy()

    # Mapeamento: nome da nossa métrica → nome da coluna no Fundamentus
    # e se deve ser normalizado de forma direta (maior = melhor)
    # ou inversa (menor = melhor)
    indicadores = {
        #  Métrica        Coluna Fundamentus         Maior é melhor?
        "dy":             ("dy",              True),  # DY alto = bom
        "roe":            ("roe",             True),  # ROE alto = bom
        "pl":             ("p/l",             False), # P/L baixo = bom
        "pvpa":           ("p/vp",            False), # P/VP baixo = bom
        # O Fundamentus não disponibiliza Dívida/EBIT diretamente.
        # Usamos Dív.Líq/Patrim. como proxy de alavancagem:
        # quanto menor, menos endividada é a empresa em relação ao patrimônio.
        "div_ebitda":     ("div.liq.patrim",  False), # Alavancagem baixa = bom
    }

    notas_calculadas = {}  # Dicionário para guardar as notas de cada indicador

    for metrica, (coluna, maior_melhor) in indicadores.items():

        # Verificamos se a coluna existe no DataFrame
        if coluna not in df_rank.columns:
            print(f"   ⚠️  Coluna '{coluna}' não encontrada — ignorando '{metrica}'")
            continue

        # Removemos valores nulos (NaN) para o cálculo
        serie = df_rank[coluna].dropna()

        if serie.empty:
            continue

        valor_min = serie.min()  # Menor valor do grupo
        valor_max = serie.max()  # Maior valor do grupo

        # Evitamos divisão por zero se todos os valores forem iguais
        if valor_max == valor_min:
            df_rank[f"nota_{metrica}"] = 0.5  # Nota neutra
            continue

        # Normalização min-max:
        # Se maior é melhor: nota = (valor - min) / (max - min)
        # Se menor é melhor: nota = (max - valor) / (max - min)
        if maior_melhor:
            df_rank[f"nota_{metrica}"] = (
                (df_rank[coluna] - valor_min) / (valor_max - valor_min)
            )
        else:
            df_rank[f"nota_{metrica}"] = (
                (valor_max - df_rank[coluna]) / (valor_max - valor_min)
            )

        notas_calculadas[metrica] = PESOS.get(metrica, 0)

    # Calculamos a nota final como média ponderada das notas individuais
    # Somamos: peso_dy * nota_dy + peso_roe * nota_roe + ...
    df_rank["nota_final"] = 0.0

    for metrica, peso in notas_calculadas.items():
        coluna_nota = f"nota_{metrica}"
        if coluna_nota in df_rank.columns:
            # Preenchemos NaN com 0 antes de multiplicar pelo peso
            df_rank["nota_final"] += df_rank[coluna_nota].fillna(0) * peso

    # Ordenamos do maior para o menor (melhores primeiro)
    df_rank = df_rank.sort_values("nota_final", ascending=False)

    # Adicionamos coluna de posição no ranking (começa em 1)
    df_rank.insert(0, "ranking", range(1, len(df_rank) + 1))

    print(f"   ✅ Ranking calculado para {len(df_rank)} ações")
    return df_rank


def buscar_historico_dividendos(tickers, anos=5):
    """
    Busca o histórico de dividendos dos últimos N anos via yfinance
    e calcula dois indicadores essenciais para carteira previdenciária:

    1. CAGR de Dividendos (Compound Annual Growth Rate):
       Taxa de crescimento anual composta dos dividendos no período.
       Fórmula: (DY_final / DY_inicial) ^ (1/anos) - 1
       Interpretação: CAGR de 10% significa que os dividendos crescem
       em média 10% ao ano — fundamental para avaliar crescimento real.

    2. Consistência de Pagamentos:
       Percentual de anos no período em que a empresa pagou dividendos.
       100% = pagou em todos os anos | 60% = pagou em 3 de 5 anos.
       Barsi prioriza empresas com histórico consistente (>= 80%).

    Parâmetros:
        tickers : lista de tickers (sem sufixo .SA)
        anos    : quantos anos de histórico buscar (padrão: 5)

    Retorna:
        DataFrame com colunas: cagr_div, consistencia_div, dy_medio_hist
    """
    import yfinance as yf
    from datetime import datetime, timedelta

    print(f"\n📅 Buscando histórico de dividendos ({anos} anos) via yfinance...")
    print(f"   (pode levar alguns segundos — uma requisição por ação)\n")

    resultados = []
    data_fim   = datetime.today()
    data_ini   = data_fim - timedelta(days=anos * 365)

    for i, ticker in enumerate(tickers, 1):
        # O yfinance usa sufixo .SA para ações brasileiras
        ticker_yf = f"{ticker}.SA"

        try:
            # Baixamos apenas os dividendos — muito mais rápido que o histórico completo
            ativo = yf.Ticker(ticker_yf)
            divs  = ativo.dividends  # Série com data e valor de cada provento

            if divs.empty:
                resultados.append({
                    "ticker":          ticker,
                    "cagr_div":        float("nan"),
                    "consistencia_div": 0.0,
                    "dy_medio_hist":   float("nan"),
                    "anos_com_div":    0,
                })
                print(f"   [{i:02d}/{len(tickers)}] {ticker:<8} — sem histórico de dividendos")
                time.sleep(0.3)
                continue

            # Filtra apenas o período de análise
            # O índice do yfinance vem com timezone — removemos para comparar datas
            divs.index = divs.index.tz_localize(None)
            divs = divs[divs.index >= pd.Timestamp(data_ini)]
            divs = divs[divs.index <= pd.Timestamp(data_fim)]

            if divs.empty:
                resultados.append({
                    "ticker":          ticker,
                    "cagr_div":        float("nan"),
                    "consistencia_div": 0.0,
                    "dy_medio_hist":   float("nan"),
                    "anos_com_div":    0,
                })
                print(f"   [{i:02d}/{len(tickers)}] {ticker:<8} — sem dividendos no período")
                time.sleep(0.3)
                continue

            # Agrupa dividendos por ano — soma todos os proventos de cada ano
            divs_anuais = divs.groupby(divs.index.year).sum()

            # Consistência: quantos anos tiveram dividendos / total de anos no período
            anos_totais   = anos
            anos_com_div  = len(divs_anuais)
            # Limitamos a 1.0 (100%) pois o yfinance pode retornar dados
            # de anos além da janela solicitada (ex: dividendos de dezembro
            # que o servidor registra em janeiro do ano seguinte).
            consistencia  = min(anos_com_div / anos_totais, 1.0)

            # CAGR de Dividendos
            # Só calculamos se tivermos pelo menos 2 anos de dados
            if len(divs_anuais) >= 2:
                div_inicial = divs_anuais.iloc[0]   # Dividendo do primeiro ano
                div_final   = divs_anuais.iloc[-1]  # Dividendo do último ano
                n_anos      = len(divs_anuais) - 1  # Intervalo entre primeiro e último

                if div_inicial > 0 and div_final > 0:
                    # Fórmula do CAGR: (valor_final / valor_inicial) ^ (1/n) - 1
                    cagr = (div_final / div_inicial) ** (1 / n_anos) - 1
                else:
                    cagr = float("nan")
            else:
                cagr = float("nan")

            # DY médio histórico: média dos dividendos anuais do período
            dy_medio = divs_anuais.mean()

            resultados.append({
                "ticker":           ticker,
                "cagr_div":         cagr,
                "consistencia_div": consistencia,
                "dy_medio_hist":    dy_medio,
                "anos_com_div":     anos_com_div,
            })

            # Formata CAGR para exibição (pode ser NaN)
            cagr_str = f"{cagr:.1%}" if not (isinstance(cagr, float) and cagr != cagr) else "N/D"
            print(f"   [{i:02d}/{len(tickers)}] {ticker:<8} — "
                  f"CAGR: {cagr_str:>8}  |  "
                  f"Consistência: {consistencia:.0%}  |  "
                  f"Anos c/ div: {anos_com_div}/{anos_totais}")

            # Pausa entre requisições para não sobrecarregar o yfinance
            time.sleep(0.5)

        except Exception as erro:
            print(f"   [{i:02d}/{len(tickers)}] {ticker:<8} — ERRO: {erro}")
            resultados.append({
                "ticker":          ticker,
                "cagr_div":        float("nan"),
                "consistencia_div": 0.0,
                "dy_medio_hist":   float("nan"),
                "anos_com_div":    0,
            })
            time.sleep(0.5)

    # Converte a lista de dicionários em DataFrame e define ticker como índice
    df_hist = pd.DataFrame(resultados).set_index("ticker")
    print(f"\n   ✅ Histórico coletado para {len(df_hist)} ações")
    return df_hist


def mesclar_historico_no_ranking(df_rank, df_historico):
    """
    Adiciona as colunas de histórico de dividendos ao DataFrame do ranking.
    Usa o índice (ticker) para fazer o cruzamento entre as duas tabelas.

    Parâmetros:
        df_rank      : DataFrame com o ranking fundamentalista
        df_historico : DataFrame retornado por buscar_historico_dividendos()

    Retorna:
        DataFrame combinado com todas as colunas
    """
    # pd.merge() une dois DataFrames por uma coluna ou índice em comum
    # how='left' mantém todas as linhas do df_rank mesmo se não houver
    # correspondência no df_historico (resultado seria NaN)
    df_completo = df_rank.merge(
        df_historico[["cagr_div", "consistencia_div", "anos_com_div"]],
        left_index=True,
        right_index=True,
        how="left"
    )
    return df_completo


def exibir_top_terminal(df_rank, top_n=TOP_N):
    """
    Exibe as melhores ações no terminal de forma formatada.

    Parâmetros:
        df_rank : DataFrame com o ranking calculado
        top_n   : Quantas ações exibir (padrão: TOP_N das configurações)
    """
    print(f"\n{'=' * 70}")
    print(f"  🏆 TOP {top_n} AÇÕES DO IBOVESPA — RANKING FUNDAMENTALISTA")
    print(f"{'=' * 70}")

    # Selecionamos as colunas mais relevantes para exibição
    # Verificamos quais existem antes de selecionar
    colunas_exibir_candidatas = {
        "ranking":         "Rank",
        "setor":           "Setor",
        "dy":              "DY",
        "p/l":             "P/L",
        "p/vp":            "P/VP",
        "roe":             "ROE",
        "cagr_div":        "CAGR Div",
        "consistencia_div":"Consist.",
        "nota_final":      "Nota",
    }

    # Filtramos apenas as colunas que existem no DataFrame
    colunas_existentes = {
        k: v for k, v in colunas_exibir_candidatas.items()
        if k in df_rank.columns
    }

    # Pegamos as top_n primeiras linhas
    df_top = df_rank.head(top_n).copy()

    # Renomeamos as colunas para exibição mais amigável
    df_exibir = df_top[list(colunas_existentes.keys())].rename(
        columns=colunas_existentes
    )

    # Formatamos os valores percentuais (DY, ROE, CAGR Div) como porcentagem
    for col_pct in ["DY", "ROE", "CAGR Div"]:
        if col_pct in df_exibir.columns:
            df_exibir[col_pct] = df_exibir[col_pct].apply(
                lambda x: f"{x*100:.1f}%" if pd.notna(x) else "N/D"
            )

    # Formatamos Consistência como percentual também
    if "Consist." in df_exibir.columns:
        df_exibir["Consist."] = df_exibir["Consist."].apply(
            lambda x: f"{x*100:.0f}%" if pd.notna(x) else "N/D"
        )

    # Formatamos a nota final com 3 casas decimais
    if "Nota" in df_exibir.columns:
        df_exibir["Nota"] = df_exibir["Nota"].apply(
            lambda x: f"{x:.3f}" if pd.notna(x) else "N/D"
        )

    # Formatamos P/L e P/VP com 2 casas decimais
    for col_dec in ["P/L", "P/VP", "Dív/EBIT"]:
        if col_dec in df_exibir.columns:
            df_exibir[col_dec] = df_exibir[col_dec].apply(
                lambda x: f"{x:.2f}" if pd.notna(x) else "N/D"
            )

    # Exibimos a tabela no terminal
    # to_string() converte o DataFrame para texto formatado
    print(df_exibir.to_string(index=True))

    print(f"\n{'=' * 70}")
    print(f"  Nota: DY e ROE maiores = melhor | P/L e P/VP menores = melhor")
    print(f"  A nota final varia de 0 (pior) a 1 (melhor)")
    print(f"{'=' * 70}\n")


def exportar_excel(df_rank):
    """
    Exporta o ranking completo para um arquivo Excel (.xlsx) formatado,
    com cores para facilitar a leitura.

    Parâmetros:
        df_rank : DataFrame com o ranking completo

    Retorna:
        Caminho do arquivo Excel gerado
    """
    # Geramos o nome do arquivo com data e hora para não sobrescrever versões anteriores
    agora = datetime.now().strftime("%Y%m%d_%H%M")
    nome_arquivo = f"ranking_acoes_ibovespa_{agora}.xlsx"
    caminho_arquivo = os.path.join(PASTA_RESULTADOS, nome_arquivo)

    print(f"\n💾 Exportando para Excel...")

    try:
        # ExcelWriter permite mais controle sobre a formatação do arquivo
        # engine='openpyxl' especifica qual biblioteca usar para gerar o .xlsx
        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:

            # --- Aba 1: Ranking Completo ---
            # Selecionamos as colunas mais relevantes para o Excel
            colunas_excel = [
                col for col in [
                    "ranking", "setor", "dy", "p/l", "p/vp", "roe",
                    "div.liq.patrim", "liq.2meses",
                    "cagr_div", "consistencia_div", "anos_com_div",
                    "nota_dy", "nota_roe", "nota_pl", "nota_pvpa",
                    "nota_div_ebitda", "nota_final"
                ]
                if col in df_rank.columns
            ]

            df_excel = df_rank[colunas_excel].copy()

            # Converte CAGR e consistência para percentual legível no Excel
            # (ex: 0.069 → 6.9  para aparecer como 6.9% ao formatar a célula)
            for col_pct in ["cagr_div", "consistencia_div"]:
                if col_pct in df_excel.columns:
                    df_excel[col_pct] = (df_excel[col_pct] * 100).round(2)

            # Exportamos o DataFrame para a primeira aba do Excel
            df_excel.to_excel(writer, sheet_name="Ranking Completo", index=True)

            # Acessamos a aba recém-criada para fazer ajustes visuais
            planilha = writer.sheets["Ranking Completo"]

            # Ajustamos a largura das colunas automaticamente
            # enumerate() retorna (posição, valor) — usamos para calcular a letra da coluna
            for idx, coluna in enumerate(df_excel.reset_index().columns):
                largura = max(len(str(coluna)) + 2, 12)  # Mínimo 12 caracteres de largura
                # get_column_letter converte número (1,2,3...) para letra (A,B,C...)
                from openpyxl.utils import get_column_letter
                planilha.column_dimensions[get_column_letter(idx + 1)].width = largura

            # --- Aba 2: Top 20 com formatação visual ---
            df_top20 = df_rank.head(20)[colunas_excel].copy()
            df_top20.to_excel(writer, sheet_name="TOP 20", index=True)

            # Aplicamos cor de fundo verde claro no cabeçalho do TOP 20
            from openpyxl.styles import PatternFill, Font, Alignment
            planilha_top = writer.sheets["TOP 20"]

            # Cor verde claro para o cabeçalho
            cor_cabecalho = PatternFill(
                start_color="C6EFCE",  # Verde claro (código hexadecimal)
                end_color="C6EFCE",
                fill_type="solid"
            )

            # Aplicamos a cor em todas as células da primeira linha (cabeçalho)
            for celula in planilha_top[1]:
                celula.fill = cor_cabecalho
                celula.font = Font(bold=True)  # Negrito no cabeçalho
                celula.alignment = Alignment(horizontal="center")

            # --- Aba 3: Metadados ---
            # Registramos informações sobre quando e como o arquivo foi gerado
            info = {
                "Informação": [
                    "Data de geração",
                    "Fonte de dados",
                    "Total de ações analisadas",
                    "Total após filtros",
                    "Filtro DY mínimo",
                    "Filtro P/L máximo",
                    "Filtro P/VP máximo",
                    "Filtro Liquidez mínima",
                    "Peso DY no ranking",
                    "Peso ROE no ranking",
                    "Peso P/L no ranking",
                    "Peso P/VP no ranking",
                    "Peso Dívida/EBITDA no ranking",
                ],
                "Valor": [
                    datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "Fundamentus.com.br + B3",
                    "Ibovespa",
                    len(df_rank),
                    f"{FILTRO_DY_MINIMO*100:.0f}%",
                    FILTRO_PL_MAXIMO,
                    FILTRO_PVPA_MAXIMO,
                    f"R$ {FILTRO_LIQUIDEZ_MINIMA:,.0f}",
                    f"{PESOS['dy']*100:.0f}%",
                    f"{PESOS['roe']*100:.0f}%",
                    f"{PESOS['pl']*100:.0f}%",
                    f"{PESOS['pvpa']*100:.0f}%",
                    f"{PESOS['div_ebitda']*100:.0f}%",
                ]
            }
            df_info = pd.DataFrame(info)
            df_info.to_excel(writer, sheet_name="Parâmetros", index=False)

        print(f"   ✅ Arquivo salvo em: {caminho_arquivo}")
        return caminho_arquivo

    except Exception as erro:
        print(f"\n❌ ERRO ao exportar Excel: {erro}")
        return None


# --- SEÇÃO 4: Execução principal ---

if __name__ == "__main__":

    # Registramos o horário de início para medir o tempo total
    inicio = datetime.now()

    print("\n" + "=" * 55)
    print("  RANKING FUNDAMENTALISTA — AÇÕES DO IBOVESPA")
    print("  Projeto B3 — Análise de Investimentos")
    print("=" * 55)

    # Passo 1: Garantir que as pastas necessárias existem
    criar_pastas()

    # Passo 2: Buscar a composição atual do Ibovespa
    tickers_ibov = buscar_composicao_ibovespa()

    # Pequena pausa entre requisições (boa prática)
    time.sleep(1)

    # Passo 3: Buscar todos os indicadores do Fundamentus
    df_todos = buscar_todos_fundamentus()

    # Se a coleta falhou, encerramos o script
    if df_todos is None:
        print("\n❌ Não foi possível continuar sem os dados. Encerrando.")
        exit(1)

    # Passo 4: Manter apenas as ações do Ibovespa
    df_ibov = filtrar_por_ibovespa(df_todos, tickers_ibov)

    # Passo 5: Aplicar filtros de qualidade
    df_filtrado = aplicar_filtros_qualidade(df_ibov)

    # Verificamos se sobrou alguma ação após os filtros
    if df_filtrado.empty:
        print("\n⚠️  Nenhuma ação passou pelos filtros. Considere flexibilizá-los.")
        print("   Dica: reduza FILTRO_DY_MINIMO ou aumente FILTRO_PL_MAXIMO.")
        exit(1)

    # Passo 6: Calcular o ranking multicritério
    df_ranking = calcular_ranking(df_filtrado)

    # Passo 7: Buscar histórico de dividendos (CAGR + consistência)
    tickers_para_historico = df_ranking.index.tolist()
    df_historico = buscar_historico_dividendos(tickers_para_historico, anos=5)

    # Passo 8: Mesclar histórico ao ranking
    df_ranking = mesclar_historico_no_ranking(df_ranking, df_historico)

    # Passo 9: Exibir o TOP N no terminal
    exibir_top_terminal(df_ranking, top_n=TOP_N)

    # Passo 10: Exportar o ranking completo para Excel
    arquivo_gerado = exportar_excel(df_ranking)

    # Calculamos o tempo total de execução
    fim = datetime.now()
    tempo_total = (fim - inicio).seconds

    # Mensagem de conclusão
    print(f"✅ Concluído em {tempo_total} segundos.")
    if arquivo_gerado:
        print(f"📂 Abra o arquivo Excel para ver o ranking completo:")
        print(f"   {os.path.abspath(arquivo_gerado)}\n")
